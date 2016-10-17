from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from pdfScript import pdfPageFunction, pdfValueFinder
import csv
from openpyxl import Workbook
from fileScript import fileDirectory, listSpecificFile, removeFiles, moveFiles
import sys
import json
from httplib2 import Http
from oauth2client.service_account import ServiceAccountCredentials
from apiclient.discovery import build

scopes = ['https://www.googleapis.com/auth/drive']

credentials = ServiceAccountCredentials.from_json_keyfile_name('service-account.json', scopes)


##Inserting access to lib directory##
sys.path.insert(0, 'lib')
#Authentification Process#
gauth = GoogleAuth()
gauth.credentials = credentials
drive = GoogleDrive(gauth)
##########################
FIND_OTN = "OTN"
FIND_SUMMARY = "Nro de Documento Interno"
FIND_FISCAL = "Detalles de la transacci"
summaryInvoicesObj = {}
fiscalInvoicesObj = {}
orderedInvoicesObj = {}
unmatchedObj = {}

#gDrive Folder Id of where invoices in pdf format are located#
folder_id = "0B2rLEVd6a2fIT3Fhek1PWWhHN28"
#gDrive Folder Id for ordered invoices#
ordered_folder_id = "0B2rLEVd6a2fIYXItOWtNR1IybTQ"

#gDrive path##
gDrive_path = '/usr/local/google/home/gilaguilar/AppEngineDevelopment/app-engine-bootcamp'

#Temporary file path#
temp_file_path = '/usr/local/google/home/gilaguilar/AppEngineDevelopment/app-engine-bootcamp/temp'
#/usr/local/google/home/gilaguilar/App Engine Development/app engine bootcamp



##Function Section############################################

##Function to identify Invoices##
def pdfInvoiceIdentifier(folder_object):
    counter = 0
    summaryCounter = 0
    fiscalCounter = 0
    while counter < len(folder_object):
        tempFile = drive.CreateFile({'id':folder_object[counter][u'id']})
        tempFile.GetContentFile(folder_object[counter][u'originalFilename'], mimetype='application/pdf')
       
        ##Open pdf file and store it in "pdfPage" variable##
        pdfPage = pdfPageFunction(folder_object[counter][u'originalFilename'])
        ##use pdfValueFinder() to detect OTN, Nro de Documento Interno y Detalles de Transacci##
        value_found_summary = pdfValueFinder(pdfPage, FIND_SUMMARY)
        value_found_fiscal = pdfValueFinder(pdfPage, FIND_FISCAL)
        value_found_otn = pdfValueFinder(pdfPage, FIND_OTN)
        print "----------------------------------------------------------------"
        print "value_found_summary: ", value_found_summary
        print "value_found_fiscal: ", value_found_fiscal
        print "value_found_otn: ", value_found_otn
        print "file name: ", folder_object[counter][u'originalFilename'];
        print "-----------------------------------------------------------------"
        ##Conditional to save summary invoice in invoice_summary object##
        if value_found_summary != -1:
            ##Save invoice in invoice_summary object##
            invoice_summary = drive.CreateFile({'id':folder_object[counter][u'id']})
            ##Debugging: print out found value##
            print "Nro de Documento Interno Found: ",pdfPage[value_found_summary+25:value_found_summary+34]
            summaryInvoicesObj["transaction"+str(summaryCounter)] ={ pdfPage[value_found_summary+25:value_found_summary+34]:invoice_summary }
            summaryCounter = summaryCounter +1
            #print summaryInvoicesObj.keys()
            #print "Documento Interno Found"
        ##Conditional to identify fiscal invoices##
        if value_found_fiscal != -1 or value_found_otn != -1:
            invoice_fiscal = drive.CreateFile({'id':folder_object[counter][u'id']})
            print "Detalles de la transaccion: ", pdfPage[value_found_fiscal+26:value_found_fiscal+35]
            #print pdfPage[value_found_fiscal+26:value_found_fiscal+35]
            print "OTN: ", pdfPage[value_found_otn+5:value_found_otn+14]
            #print pdfPage[value_found_otn+5:value_found_otn+13]
            if value_found_fiscal != -1:
                fiscalInvoicesObj["transaction"+str(fiscalCounter)] = {pdfPage[value_found_fiscal+26:value_found_fiscal+35]:invoice_fiscal}
            else:
                fiscalInvoicesObj["transaction"+str(fiscalCounter)] = {pdfPage[value_found_otn+5:value_found_otn+14]:invoice_fiscal}
            #fiscalInvoicesObj["transaction"+str(fiscalCounter)] = {pdfPage[value_found_fiscal+27:value_found_fiscal+37]:invoice_fiscal}
            fiscalCounter = fiscalCounter + 1 
            #print fiscalInvoicesObj.keys()
            print "Detalles de la Transaccion Found" 
        counter = counter + 1
#Function that lists Files on specified Folder Identification#
def list_folder(folder_id):
    # folder_id: GoogleDriveFile['id']###########################################################
    _q = {'q': "'{}' in parents and trashed=false".format(folder_id)}
    return drive.ListFile(_q).GetList()

#Fuction that creates Google Drive Folder in root "My Drive" by providing string with "folder_name"#
def create_gFolder(folder_name):
    folderDrive = drive.CreateFile({'title':folder_name, 'mimeType':'application/vnd.google-apps.folder'})
    folderDrive.Upload()

#Function that uploads files into specified gFolder using ID#
def upload_file(gFolderId, filePath):
    f = drive.CreateFile({"parents": [{"kind": "drive#fileLink", "id": gFolderId}]})
    f.SetContentFile(filePath)
    f.Upload()

#Function to create CSV file#
def create_excel(invoiceObj):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Summary Invoice"
    ws['B1'] = "Summary Invoice Download"
    ws['C1'] = "Mexican Fiscal Invoice"
    ws['D1'] = "Mexican Fiscal Invoice Download"
    ws['E1'] = "Client Number"
    counter = 0 
    while counter < len(invoiceObj):
        transId = invoiceObj["transaction"+str(counter)]["summaryInvoice"].keys()
        ws['A'+str(counter+2)] = invoiceObj["transaction"+str(counter)]["summaryInvoice"][transId[0]][u'originalFilename']
        ws['B'+str(counter+2)] = invoiceObj["transaction"+str(counter)]["summaryInvoice"][transId[0]][u'webContentLink']
        ws['C'+str(counter+2)] = invoiceObj["transaction"+str(counter)]["fiscalInvoice"][transId[0]][u'originalFilename']
        ws['D'+str(counter+2)] = invoiceObj["transaction"+str(counter)]["fiscalInvoice"][transId[0]][u'webContentLink']
       # ws['E'+str(counter+2)] = invoiceObj["client_name"]
        counter = counter + 1
    wb.save("orderedInvoices.xlsx")
##Function Section##############################################
    

# Auto-iterate through all files that matches this query
file_list = drive.ListFile({'q': "'root' in parents"}).GetList()
#for file1 in file_list:
  #print('title: %s, id: %s' % (file1['title'], file1['id']))

#gDrive Folder Id of where invoices in pdf format are located#
folder_id = "0B2rLEVd6a2fIT3Fhek1PWWhHN28"

#Upload excel file with organized invoices# 
file_path = "/usr/local/google/home/gilaguilar/AppEngineDevelopment/app-engine-bootcamp/orderedInvoices.xlsx"
#Example of to download using file Id#
fileDownload = drive.CreateFile({'id':folder_id})


#Use list_folder() method to download gDrive folder identified by ID#
folderObj= list_folder(folder_id)


##Identify invoices into summary and fiscal objects##
pdfInvoiceIdentifier(folderObj)


##Section for Debugging## 
#print "Invoices Summary-------------------------------------------------------------------------------------------"
#print summaryInvoicesObj
#print "-----------------------------------------------------------------------------------------------------------"


#print "Invoices Fiscales------------------------------------------------------------------------------------------"
#print fiscalInvoicesObj
#print "-----------------------------------------------------------------------------------------------------------"

summaryCounter = 0
fiscalCounter = 0
macroCounter = 0

#print "Number of Summary Invoices--------------------------", len(summaryInvoicesObj)
#print "Number of Fiscal Invoices-------------------------- ", len(fiscalInvoicesObj)

while summaryCounter < len(summaryInvoicesObj):
    #print "summary Counter:",summaryCounter
    fiscalCounter = 0
    while fiscalCounter < len(fiscalInvoicesObj):
        #print "fiscal counter:", fiscalCounter
        if summaryInvoicesObj["transaction"+str(summaryCounter)].keys() == fiscalInvoicesObj["transaction"+str(fiscalCounter)].keys():
            print "Match......................................................................"
            orderedInvoicesObj["transaction"+str(macroCounter)]= {"summaryInvoice":summaryInvoicesObj["transaction"+str(summaryCounter)], "fiscalInvoice":fiscalInvoicesObj["transaction"+str(fiscalCounter)]}
            print "macroCounter:", macroCounter
            macroCounter = macroCounter + 1
            fiscalCounter = 0
            break 
        fiscalCounter = fiscalCounter + 1
    unmatchedObj["invoice"+str(summaryCounter)] = summaryInvoicesObj["transaction"+str(summaryCounter)]
    summaryCounter = summaryCounter + 1

print "Ordered Invoices Object---------------------------------------------------------------------------------"
print orderedInvoicesObj 
print "--------------------------------------------------------------------------------------------------------"    

print "Unmatched Inovices Object-------------------------------------------------------------------------------"
print unmatchedObj
print "--------------------------------------------------------------------------------------------------------"


           
create_excel(orderedInvoicesObj)

##Upload excel file in folder in gDrive##
upload_file(ordered_folder_id, file_path)


#pdf_files_dir = listSpecificFile(gDrive_path, '.pdf')
    
#counter = 0
#while counter < len(pdf_files_dir):
#    moveFiles(pdf_files_dir[counter],temp_file_path)
#    counter = counter + 1
#
#removeFiles (temp_file_path+'/*')
